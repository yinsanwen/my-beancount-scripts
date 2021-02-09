from beancount.core import data
from beancount.core.data import Transaction
from openpyxl import load_workbook

from modules.imports.base import Base


def get_base_account(account):
    pass


def get_income_account(category):
    pass


def get_expenses_account(category):
    pass


class CMBDebit(Base):

    def __init__(self, filename):
        workbook = load_workbook(filename)
        sheet_names = workbook.sheetnames
        print(sheet_names)
        self.sheet = workbook.active
        for row in self.sheet:
            print(row[0].value, end=' |')
            print(row[1].value, end=' |')
            print(row[2].value, end=' |')
            print(row[3].value, end=' |')
            print(row[4].value, end=' |')
            print(row[5].value, end=' |')
            print(row[6].value, end=' |')
            print(row[7].value, end=' |')
            print(row[8].value, end=' |')
            print(row[9].value, end=' |')
            print('\n')

    def parse(self):
        transactions = []
        for row in self.sheet:
            if row[0].value == '交易时间': continue
            transactions.append(self.gen_transaction(row))
        return transactions

    def gen_transaction(self, row):
        print(self.tradetime(row))
        trade_time = row[0]

        trade_time = row[0].value
        trade_type = row[1].value
        trade_partner = row[2].value
        product_description = row[3].value
        income_outcome = row[4].value
        category = row[5].value
        amount = str(row[6].value)
        project = row[7].value
        account = row[8].value
        comment = row[9].value

        tags = set()
        if project is not None:
            tags.add(project)
        if comment is not None and comment != '/':
            product_description += ' @' + comment

        meta = {}
        meta['trade_time'] = str(trade_time)

        Transaction()

        entry = Transaction(meta, trade_time.strftime('%Y-%m-%d'),
                            '*', trade_partner, product_description,
                            tags,
                            data.EMPTY_SET, [])

        # 生成基础posting
        base_post_amount = amount
        if income_outcome == '支出' or income_outcome == '转出':
            base_post_amount = '-' + amount
        data.create_simple_posting(entry, get_base_account(account), base_post_amount, 'CNY')

        if income_outcome == '收入':
            data.create_simple_posting(entry, get_income_account(category), '-' + amount, 'CNY')
        if income_outcome == '支出':
            data.create_simple_posting(entry, get_expenses_account(category), amount, 'CNY')
        if income_outcome == '转出':
            data.create_simple_posting(entry, get_base_account(category), amount, 'CNY')
        if income_outcome == '转入':
            data.create_simple_posting(entry, get_base_account(category), '-' + amount, 'CNY')
        if income_outcome == '退款':
            data.create_simple_posting(entry, get_expenses_account(category), '-' + amount, 'CNY')

        print(entry)

        return entry

    def tradetime(self, row):
        tradetime = row[0].value + row[1].value
        return tradetime


if __name__ == "__main__":
    filename = '/Users/yinsanwen/个人材料/2020年盘点/招商银行0612_20200101_20201231.xlsx'
    cmbdebit = CMBDebit(filename)
    cmbdebit.parse()
